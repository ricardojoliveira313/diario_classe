import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook('DIARIO_CLASSE_2026.xlsx', data_only=True)
mes_map = {'FEVEREIRO':2,'MARCO':3,'ABRIL':4}

alunos = {}
for sn in wb.sheetnames:
    ws = wb[sn]
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = str(row[3] or '').strip()
        if not nome or len(nome) < 3: continue
        mes_col = str(row[4] or '').strip().upper().replace('Ç','C').replace('Ã','A').replace('Í','I')
        mes = mes_map.get(mes_col, 0)
        if mes == 0: continue
        turma = str(row[1] or '').strip()
        prof = str(row[0] or '').strip()
        nasc = str(row[9] or '').strip()
        freq = row[5]
        freq_txt = str(freq).strip() if isinstance(freq, str) else ''
        freq_num = int(freq) if isinstance(freq, (int, float)) else 0
        situ = str(row[6] or 'ATIVO').strip()
        dt_mov = str(row[12] or '').strip()
        defi = str(row[13] or '').strip()
        bolsa = str(row[14] or '').strip().upper()
        key = nome.upper() + '|' + nasc
        if key not in alunos:
            alunos[key] = {'nome':nome,'nasc':nasc,'turma':turma,'prof':prof,'defi':defi,'bolsa':bolsa=='SIM','meses':{}}
        if mes not in alunos[key]['meses']:
            alunos[key]['meses'][mes] = []
        alunos[key]['meses'][mes].append({'freq_txt':freq_txt,'freq_num':freq_num,'situ':situ,'dt_mov':dt_mov})

for mes_num, mes_nome in [(2,'FEVEREIRO'),(3,'MARCO'),(4,'ABRIL')]:
    print(f'========== {mes_nome} ==========')
    presentes = [a for a in alunos.values() if mes_num in a['meses']]
    ativos = [a for a in presentes if any(m['situ']=='ATIVO' for m in a['meses'][mes_num])]
    bf = [a for a in presentes if a['bolsa']]
    defic = [a for a in presentes if a['defi']]
    print(f'Total alunos: {len(presentes)}')
    print(f'Ativos:       {len(ativos)}')
    print(f'Bolsa Família: {len(bf)}')
    print(f'Deficiência:  {len(defic)}')

    movs = []
    for a in presentes:
        for m in a['meses'][mes_num]:
            ft = m['freq_txt'].upper().replace('À','A').replace('Á','A').replace('Ã','A').replace('Ç','C').replace('Ê','E').replace('É','E').replace('Í','I').replace('Ó','O').replace('Ô','O').replace('Õ','O').replace('Ú','U')
            if ft and ft not in ('NAO HA FALTAS NO MES','NAO HA FALTAS NO','','0','0.0'):
                movs.append((a['nome'],a['turma'],m['freq_txt'],m['dt_mov'],m['situ']))
    if movs:
        print(f'\nMovimentacoes ({len(set(movs))}):')
        for n,t,ft,dtm,st in sorted(set(movs), key=lambda x: x[3] or ''):
            print(f'  {n} | {t} | {ft} | data: {dtm} | situ: {st}')
    else:
        print('Nenhuma movimentacao')
    print()
