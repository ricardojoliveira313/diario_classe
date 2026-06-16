#!/usr/bin/env python3
"""Importa dados dos PDFs da SED + Excel para o Diario de Classe.

Uso:
    python import_sed.py --pdf-dir <pasta_com_pdfs> --excel <caminho_excel1> [--excel <caminho_excel2> ...]
    
Imprime JSON no stdout.
    
Pode passar varios Excels com --excel para cobrir diferentes turmas.
"""

import argparse
import json
import os
import re
import sys
import unicodedata

import pdfplumber
from openpyxl import load_workbook


# ─── Utilitarios ─────────────────────────────────────────────────────────────

def tirar_acentos(texto: str) -> str:
    return unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode()

def limpar_nome(nome: str) -> str:
    return re.sub(r'\s+', ' ', nome).strip().upper()


def extrair_ra(ra: str) -> str:
    return re.sub(r'\D', '', ra).lstrip('0')


def normalizar_data(data: str) -> str:
    data = data.strip()
    if not data:
        return ''
    if re.match(r'^\d{2}/\d{2}/\d{4}$', data):
        return data
    partes = re.findall(r'\d+', data)
    if len(partes) == 3:
        return f'{partes[0]}/{partes[1]}/{partes[2]}'
    return data


# ─── Parser de PDF ───────────────────────────────────────────────────────────

SITUACOES = {'ATIVO', 'TRAN', 'REMA', 'BXTR', 'ABAN', 'NCOM', 'TRANSFERIDO', 'ABANDONO'}


def detectar_formato_colunas(header: list[str], primeira_linha: list[str]) -> dict:
    """Detecta o indice de cada coluna no PDF."""
    col = {}
    for i, h in enumerate(header):
        h_upper = h.upper().replace('\n', ' ').strip()
        if 'NOME' in h_upper:
            col['nome'] = i
        elif h_upper == 'RA' or h_upper.startswith('RA'):
            col['ra'] = i
        elif 'DIG' in h_upper:
            col['digito'] = i
        elif 'UF' in h_upper:
            col['uf'] = i
        elif 'NASCIMENTO' in h_upper:
            col['nascimento'] = i
        elif 'SITUACAO' in h_upper or 'SITUAÇÃO' in h_upper:
            col['situacao'] = i
        elif 'MOVIMENTACAO' in h_upper or 'MOVIMENTAÇÃO' in h_upper:
            col['data_mov'] = i
        elif 'DEFICIENCIA' in h_upper or 'DEFICIÊNCIA' in h_upper:
            col['deficiencia'] = i
        elif 'CONDICOES' in h_upper or 'CONDIÇÕES' in h_upper or 'CONDIÇOES' in h_upper:
            col['condicoes'] = i
        elif 'TRANSTORNO' in h_upper:
            col['transtornos'] = i
        elif 'SERIE' in h_upper or 'SÉRIE' in h_upper:
            col['serie'] = i
        elif h_upper == 'NR' or h_upper == 'N§' or h_upper == 'Nº':
            col['nr'] = i

    if 'situacao' not in col or 'data_mov' not in col:
        if len(primeira_linha) >= 9:
            sétimo = str(primeira_linha[7] or '').strip().upper()
            oitavo = str(primeira_linha[8] or '').strip().upper()
            if sétimo in SITUACOES or any(s in sétimo for s in SITUACOES):
                col['situacao'] = 7
                col['data_mov'] = 8
            elif oitavo in SITUACOES or any(s in oitavo for s in SITUACOES):
                col['data_mov'] = 7
                col['situacao'] = 8

    # Validar com dados reais
    if col.get('nascimento') is not None:
        val_nasc = str(primeira_linha[col['nascimento']] or '').strip()
        if not re.match(r'\d{2}/\d{2}/\d{4}', val_nasc):
            for i in range(len(primeira_linha)):
                val = str(primeira_linha[i] or '').strip()
                if re.match(r'\d{2}/\d{2}/\d{4}', val) and i not in (col.get('data_mov'), col.get('nr')):
                    col['nascimento'] = i
                    break

    if col.get('situacao') is not None:
        val_sit = str(primeira_linha[col['situacao']] or '').strip().upper()
        if val_sit not in SITUACOES:
            for i in range(len(primeira_linha)):
                val = str(primeira_linha[i] or '').strip().upper()
                if val in SITUACOES:
                    col['situacao'] = i
                    break

    if col.get('data_mov') is not None:
        val_mov = str(primeira_linha[col['data_mov']] or '').strip()
        if not re.match(r'\d{2}/\d{2}/\d{4}', val_mov.replace('\n', '')):
            for i in range(len(primeira_linha)):
                val = str(primeira_linha[i] or '').strip()
                if re.match(r'\d{2}/\d{2}/\d{4}', val.replace('\n', '')) and i != col.get('nascimento'):
                    col['data_mov'] = i
                    break

    if 'deficiencia' not in col:
        col_ignorar = {col.get('situacao'), col.get('data_mov'), col.get('condicoes'), col.get('transtornos'), col.get('nascimento'), col.get('serie'), col.get('nr'), col.get('digito'), col.get('uf'), col.get('ra'), col.get('nome')}
        for i in range(len(primeira_linha) - 1, -1, -1):
            val = str(primeira_linha[i] or '').strip()
            if val and i not in col_ignorar:
                if col.get('nome') and i > col['nome']:
                    col['deficiencia'] = i
                    break

    return col


def parse_pdf(caminho: str) -> list[dict]:
    """Extrai alunos de um PDF da SED."""
    alunos = []
    turma_atual = {}

    with pdfplumber.open(caminho) as pdf:
        for pagina in pdf.pages:
            textos = pagina.extract_text() or ''

            turma_info = {}
            for linha in textos.split('\n'):
                linha = linha.strip()
                if linha.startswith('Turma:'):
                    turma_info['turma'] = linha.replace('Turma:', '').strip()
                if 'NR. Classe:' in linha:
                    m = re.search(r'NR\. Classe:\s*(\d+)', linha)
                    if m:
                        turma_info['nr_classe'] = m.group(1)
                if 'Tipo Ensino:' in linha:
                    m = re.search(r'Tipo Ensino:\s*(.+)', linha)
                    if m:
                        turma_info['tipo_ensino'] = m.group(1).strip()
                if linha.startswith('Ano Letivo:'):
                    m = re.search(r'Ano Letivo:\s*(\d{4})', linha)
                    if m:
                        turma_info['ano_letivo'] = m.group(1)

            tabelas = pagina.extract_tables()
            if not tabelas:
                continue

            for tabela in tabelas:
                if len(tabela) < 2:
                    continue

                header = tabela[0]
                header_str = ' '.join(str(h or '') for h in header).upper().strip()

                if 'NOME' not in header_str and 'ALUNO' not in header_str:
                    continue

                col = {}
                for i, linha_dados in enumerate(tabela[1:], 1):
                    if not any(cel and str(cel).strip() for cel in linha_dados):
                        continue

                    if not col:
                        col = detectar_formato_colunas(header, linha_dados)

                    nome_raw = str(linha_dados[col.get('nome', 2)] or '').strip() if col.get('nome') is not None else ''
                    ra_raw = str(linha_dados[col.get('ra', 3)] or '').strip() if col.get('ra') is not None else ''
                    nasc_raw = str(linha_dados[col.get('nascimento', 6)] or '').strip() if col.get('nascimento') is not None else ''

                    if not nome_raw or not ra_raw:
                        continue

                    nome = limpar_nome(nome_raw)
                    ra = extrair_ra(ra_raw)
                    nascimento = normalizar_data(nasc_raw)

                    situacao_idx = col.get('situacao')
                    data_mov_idx = col.get('data_mov')
                    deficiencia_idx = col.get('deficiencia')

                    situacao = limpar_nome(str(linha_dados[situacao_idx] or '')) if situacao_idx is not None else ''
                    data_mov = normalizar_data(str(linha_dados[data_mov_idx] or '')) if data_mov_idx is not None else ''
                    deficiencia = limpar_nome(str(linha_dados[deficiencia_idx] or '')) if deficiencia_idx is not None else ''

                    nr_aluno = str(linha_dados[col.get('nr', 1)] or '').strip() if col.get('nr') is not None else ''

                    aluno = {
                        'nome': nome,
                        'ra': ra,
                        'digito': str(linha_dados[col.get('digito', 4)] or '').strip() if col.get('digito') is not None else '',
                        'uf': str(linha_dados[col.get('uf', 5)] or '').strip() if col.get('uf') is not None else '',
                        'nascimento': nascimento,
                        'situacao': situacao,
                        'dataMovimentacao': data_mov,
                        'deficiencia': deficiencia,
                        'turma': turma_info.get('turma', ''),
                        'tipoEnsino': turma_info.get('tipo_ensino', ''),
                        'nrClasse': turma_info.get('nr_classe', ''),
                        'numeroAluno': nr_aluno,
                    }
                    alunos.append(aluno)

    return alunos


# ─── Leitor de Excel ─────────────────────────────────────────────────────────

def ler_excel(caminho: str) -> list[dict]:
    """Le planilha da SED e retorna alunos com Data Inicio/Fim Matricula."""
    wb = load_workbook(caminho, data_only=True)
    ws = wb.active

    header = [str(c.value or '').strip().upper() for c in ws[1]]
    header_sem_acc = [tirar_acentos(h) for h in header]

    col_idx = {}
    for i, (h, h_plain) in enumerate(zip(header, header_sem_acc)):
        if 'NOME' in h and 'ALUNO' in h:
            col_idx['nome'] = i
        elif h == 'RA' or h_plain == 'RA':
            col_idx['ra'] = i
        elif 'DIG' in h or h_plain == 'DIG' or 'DIG' in h_plain:
            col_idx['digito'] = i
        elif 'UF' in h or 'UF' in h_plain:
            col_idx['uf'] = i
        elif 'NASCIMENTO' in h_plain:
            col_idx['nascimento'] = i
        elif 'INICIO' in h_plain and 'MATRICULA' in h_plain:
            col_idx['data_inicio_matricula'] = i
        elif 'FIM' in h_plain and 'MATRICULA' in h_plain:
            col_idx['data_fim_matricula'] = i
        elif 'SITUACAO' in h_plain:
            col_idx['situacao'] = i
        elif 'MOVIMENTACAO' in h_plain:
            col_idx['data_movimentacao'] = i
        elif 'CONDICOES' in h_plain or 'CONDIÇÕES' in h_plain:
            col_idx['condicoes'] = i
        elif 'DEFICIENCIA' in h_plain or 'DEFICIÊNCIA' in h_plain:
            col_idx['deficiencia'] = i

    alunos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        nome_raw = str(row[col_idx.get('nome', 4)] or '').strip() if col_idx.get('nome') is not None else ''
        ra_raw = str(row[col_idx.get('ra', 5)] or '').strip() if col_idx.get('ra') is not None else ''
        if not nome_raw or not ra_raw:
            continue

        nr_classe = str(row[0] or '').strip() if len(row) > 0 else ''
        nasc_raw = str(row[col_idx.get('nascimento', 8)] or '').strip() if col_idx.get('nascimento') is not None else ''
        data_ini = str(row[col_idx.get('data_inicio_matricula', 9)] or '').strip() if col_idx.get('data_inicio_matricula') is not None else ''
        data_fim = str(row[col_idx.get('data_fim_matricula', 10)] or '').strip() if col_idx.get('data_fim_matricula') is not None else ''

        condicoes_raw = str(row[col_idx.get('condicoes', 13)] or '').strip() if col_idx.get('condicoes') is not None else ''
        def_raw = str(row[col_idx.get('deficiencia', 14)] or '').strip() if col_idx.get('deficiencia') is not None else ''
        deficiencia_excel = limpar_nome(condicoes_raw or def_raw)

        aluno = {
            'nome': limpar_nome(nome_raw),
            'ra': extrair_ra(ra_raw),
            'nascimento': normalizar_data(nasc_raw),
            'dataInicioMatric': normalizar_data(data_ini),
            'dataFimMatric': normalizar_data(data_fim),
            'deficiencia': deficiencia_excel,
            'nrClasse': nr_classe,
        }
        alunos.append(aluno)

    return alunos


# ─── Matching ────────────────────────────────────────────────────────────────

def fazer_matching(alunos_pdf: list[dict], alunos_excel: list[dict]) -> list[dict]:
    """Cruza dados do PDF com Excel por Nome + RA + Data Nascimento + Turma."""
    excel_index: dict[tuple[str, str, str, str], dict] = {}
    for a in alunos_excel:
        chave = (a['nome'], a['ra'], a['nascimento'], a.get('nrClasse', ''))
        excel_index[chave] = a

    importados = []
    sem_match = []
    for aluno in alunos_pdf:
        chave = (aluno['nome'], aluno['ra'], aluno['nascimento'], aluno.get('nrClasse', ''))
        match = excel_index.get(chave)
        if match:
            aluno['dataInicioMatric'] = match.get('dataInicioMatric', '')
            aluno['dataFimMatric'] = match.get('dataFimMatric', '')
        else:
            chave_sem_turma = (aluno['nome'], aluno['ra'], aluno['nascimento'])
            match_fallback = None
            for k, v in excel_index.items():
                if (k[0], k[1], k[2]) == chave_sem_turma:
                    match_fallback = v
                    break
            if match_fallback:
                aluno['dataInicioMatric'] = match_fallback.get('dataInicioMatric', '')
                aluno['dataFimMatric'] = match_fallback.get('dataFimMatric', '')
            else:
                aluno['dataInicioMatric'] = ''
                aluno['dataFimMatric'] = ''
                sem_match.append(aluno)

        importados.append(aluno)

    return importados, sem_match


# ─── Export JSON ─────────────────────────────────────────────────────────────

def exportar_json(alunos: list[dict], caminho: str):
    with open(caminho, 'w', encoding='utf-8') as f:
        json.dump({'alunos': alunos}, f, ensure_ascii=False, indent=2)


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Importa dados da SED')
    parser.add_argument('--pdf-dir', required=True, help='Pasta com os PDFs')
    parser.add_argument('--excel', required=True, action='append', help='Caminho do Excel SED (pode repetir)')
    parser.add_argument('--output', help='Arquivo JSON de saida (opcional)')
    args = parser.parse_args()

    todos_alunos = []
    pdf_arquivos = sorted([
        os.path.join(args.pdf_dir, f)
        for f in os.listdir(args.pdf_dir)
        if f.lower().endswith('.pdf') and f.lower() != 'bolsa_familia.pdf'
    ])

    if not pdf_arquivos:
        print('Nenhum PDF encontrado em', args.pdf_dir, file=sys.stderr)
        sys.exit(1)

    print(f'PDFs encontrados: {len(pdf_arquivos)}', file=sys.stderr)
    for pdf_path in pdf_arquivos:
        nome_pdf = os.path.basename(pdf_path)
        print(f'  Parseando {nome_pdf}...', file=sys.stderr)
        try:
            alunos = parse_pdf(pdf_path)
            print(f'    -> {len(alunos)} alunos extraidos', file=sys.stderr)
            todos_alunos.extend(alunos)
        except Exception as e:
            print(f'    -> ERRO: {e}', file=sys.stderr)

    print(f'\nTotal de alunos nos PDFs: {len(todos_alunos)}', file=sys.stderr)

    excel_alunos = []
    for excel_path in args.excel:
        print(f'Lendo Excel: {excel_path}...', file=sys.stderr)
        alunos = ler_excel(excel_path)
        print(f'  -> {len(alunos)} alunos', file=sys.stderr)
        excel_alunos.extend(alunos)
    print(f'Total de alunos nos Excels: {len(excel_alunos)}', file=sys.stderr)

    importados, sem_match = fazer_matching(todos_alunos, excel_alunos)
    print(f'\nAlunos com match: {len(importados) - len(sem_match)}', file=sys.stderr)
    print(f'Alunos SEM match (data vazia): {len(sem_match)}', file=sys.stderr)

    for a in sem_match[:5]:
        print(f'  SEM MATCH: {a["nome"]} | RA: {a["ra"]} | Nasc: {a["nascimento"]}', file=sys.stderr)

    if args.output:
        exportar_json(importados, args.output)
        print(f'\nJSON exportado para: {args.output}', file=sys.stderr)

    print(json.dumps({'alunos': importados, 'total': len(importados), 'semMatch': len(sem_match)}, ensure_ascii=False))


if __name__ == '__main__':
    main()
